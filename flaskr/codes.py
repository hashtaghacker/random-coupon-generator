import openpyxl
import secrets
from flask import (
    Blueprint, render_template
)
from flaskr.auth import login_required

bp = Blueprint('blog', __name__)


@bp.route('/codes')
@login_required
def index():
    number = secrets.randbelow(1000000)
    return render_template('blog/index.html', number=number)


def download_codes():
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.title = 'Sheet #1'

    # Generate data
    sno = []
    sno.extend(range(100))
    activation_code = []

    for i in sno:
        activation_code.append(str(secrets.randbits(32))[:6]),

    # Add titles in the first row of each column
    sheet.cell(row=1, column=1).value = 'S. NO'
    sheet.cell(row=1, column=2).value = 'Activation Code'
    sheet.cell(row=1, column=3).value = 'Jira Tickets'
    sheet.cell(row=1, column=4).value = 'Issued for'

    # Loop to set the activation codes in each cells

    for inputs in range(0, len(sno)):
        sheet.cell(row=inputs + 2, column=1).value = sno[inputs]
        sheet.cell(row=inputs + 2, column=2).value = activation_code[inputs]

    # Save the activation codes
    save_file = wb.save('Activation_Codes.xlsx')

    return render_template('blog/index.html', save_file=save_file)
